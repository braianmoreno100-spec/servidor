"""
excel_export.py — Reportes Excel Lean + endpoints JSON para el dashboard
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta, date
from io import BytesIO
import calendar
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

from ..database import get_db
from ..models import Turno, RegistroProduccion, Parada, Desperdicio, Relevo, Orden, Producto

router = APIRouter(tags=["reportes"])

# ─────────────────────────────────────────────
# CONSTANTES LEAN
# ─────────────────────────────────────────────
MINUTOS_TURNO = 480
HORAS_TURNO   = 8

# ─────────────────────────────────────────────
# PALETA DE COLORES
# ─────────────────────────────────────────────
COLOR_HEADER_DARK   = "1F3864"
COLOR_HEADER_MID    = "2E75B6"
COLOR_HEADER_LIGHT  = "D6E4F0"
COLOR_VERDE_GOOD    = "E2EFDA"
COLOR_ROJO_BAD      = "FFDFD6"
COLOR_AMARILLO_WARN = "FFF2CC"
COLOR_GRIS_ROW      = "F2F2F2"
COLOR_LEAN_ACCENT   = "375623"
COLOR_TITULO_SHEET  = "16365C"

# ─────────────────────────────────────────────
# kW FIJOS POR MÁQUINA
# ─────────────────────────────────────────────
KW_MAQUINA = {
    ("inyeccion", "1"): 18.5,
    ("inyeccion", "2"): 15.0,
    ("inyeccion", "3"): 11.0,
    ("inyeccion", "4"): 22.0,
    ("inyeccion", "5"): 13.0,
    ("inyeccion", "6"): 11.0,
    ("inyeccion", "7"): 13.0,
    ("soplado",   "1"): 45.1,
    ("linea",     "1"):  5.0,  # pendiente dato real
    ("linea",     "2"):  5.0,  # pendiente dato real
    ("acondicionamiento", "1"): 3.0,  # pendiente dato real
    ("acondicionamiento", "2"): 3.0,  # pendiente dato real
}

# ─────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border_thin():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _header_cell(ws, row, col, value, bg=COLOR_HEADER_MID, fg="FFFFFF", size=10, bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, color=fg, size=size)
    cell.alignment = _center()
    cell.border = _border_thin()
    return cell

def _data_cell(ws, row, col, value, bg=None, bold=False, align="center", num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = _fill(bg)
    cell.font = _font(bold=bold, size=10)
    cell.alignment = _center() if align == "center" else _left()
    cell.border = _border_thin()
    if num_format:
        cell.number_format = num_format
    return cell

def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def _color_oee(valor: float) -> str:
    if valor >= 85:
        return COLOR_VERDE_GOOD
    elif valor >= 60:
        return COLOR_AMARILLO_WARN
    return COLOR_ROJO_BAD

# ─────────────────────────────────────────────
# CÁLCULOS LEAN
# ─────────────────────────────────────────────
def calcular_metricas_lean(turno: Turno, orden: Orden) -> dict:
    total_producido   = sum(r.cantidad for r in turno.registros_produccion)
    total_desperdicio = sum(d.cantidad for d in turno.desperdicios)
    piezas_buenas     = total_producido - total_desperdicio

    min_paradas_no_prog = sum(p.minutos for p in turno.paradas if not p.programada)
    min_paradas_prog    = sum(p.minutos for p in turno.paradas if p.programada)
    total_paradas_min   = min_paradas_no_prog + min_paradas_prog

    tiempo_disponible = MINUTOS_TURNO - min_paradas_prog
    tiempo_operativo  = tiempo_disponible - min_paradas_no_prog

    disponibilidad = (tiempo_operativo / tiempo_disponible * 100) if tiempo_disponible > 0 else 0

    ciclos_min         = (orden.ciclos / 60) if orden.ciclos else 0
    cavidades          = orden.cavidades or 1
    produccion_teorica = ciclos_min * cavidades * tiempo_operativo
    rendimiento        = (total_producido / produccion_teorica * 100) if produccion_teorica > 0 else 0
    rendimiento        = min(rendimiento, 100)

    calidad = (piezas_buenas / total_producido * 100) if total_producido > 0 else 0
    oee     = (disponibilidad * rendimiento * calidad) / 10000

    num_paradas = len([p for p in turno.paradas if not p.programada])
    mttr = (min_paradas_no_prog / num_paradas) if num_paradas > 0 else 0
    mtbf = ((tiempo_operativo - min_paradas_no_prog) / num_paradas) if num_paradas > 0 else tiempo_operativo

    horas_operativas = tiempo_operativo / 60
    uph_real    = (total_producido / horas_operativas) if horas_operativas > 0 else 0
    uph_teorica = (produccion_teorica / horas_operativas) if horas_operativas > 0 else 0

    return {
        "total_producido":     total_producido,
        "total_desperdicio":   total_desperdicio,
        "piezas_buenas":       piezas_buenas,
        "produccion_teorica":  round(produccion_teorica),
        "min_paradas_prog":    min_paradas_prog,
        "min_paradas_no_prog": min_paradas_no_prog,
        "total_paradas_min":   total_paradas_min,
        "tiempo_disponible":   tiempo_disponible,
        "tiempo_operativo":    tiempo_operativo,
        "disponibilidad":      round(disponibilidad, 1),
        "rendimiento":         round(rendimiento, 1),
        "calidad":             round(calidad, 1),
        "oee":                 round(oee, 1),
        "mttr":                round(mttr, 1),
        "mtbf":                round(mtbf, 1),
        "uph_real":            round(uph_real, 1),
        "uph_teorica":         round(uph_teorica, 1),
        "num_paradas":         len(turno.paradas),
        "num_paradas_no_prog": num_paradas,
        "tasa_desperdicio":    round((total_desperdicio / total_producido * 100) if total_producido > 0 else 0, 2),
    }

# ─────────────────────────────────────────────
# HOJAS EXCEL
# ─────────────────────────────────────────────
def _hoja_resumen_lean(wb, turnos, ordenes_map, titulo):
    ws = wb.create_sheet("Resumen Lean")
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = titulo
    c.fill = _fill(COLOR_TITULO_SHEET)
    c.font = _font(bold=True, color="FFFFFF", size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:R2")
    c = ws["A2"]
    c.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Metodología: Lean Manufacturing"
    c.fill = _fill(COLOR_HEADER_LIGHT)
    c.font = _font(color=COLOR_LEAN_ACCENT, size=10)
    c.alignment = _center()

    headers = [
        "Orden", "Producto", "Tipo Máq.", "Máq. #", "Fecha", "Turno", "Empleado",
        "Prod. Real", "Prod. Teórica", "Piezas Buenas", "Desperdicio",
        "T. Paradas\n(min)", "Paradas\nNo Prog.",
        "Disponib.\n%", "Rendim.\n%", "Calidad\n%", "OEE\n%", "MTTR\n(min)"
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 3, col, h, bg=COLOR_HEADER_DARK, size=9)
    ws.row_dimensions[3].height = 35

    for i, turno in enumerate(turnos, 4):
        orden = ordenes_map.get(turno.orden_id)
        if not orden:
            continue
        m  = calcular_metricas_lean(turno, orden)
        bg = COLOR_GRIS_ROW if i % 2 == 0 else None

        _data_cell(ws, i, 1,  orden.numero_orden,        bg=bg, align="left")
        _data_cell(ws, i, 2,  orden.descripcion_producto, bg=bg, align="left")
        _data_cell(ws, i, 3,  orden.tipo_maquina,         bg=bg)
        _data_cell(ws, i, 4,  orden.numero_maquina,       bg=bg)
        _data_cell(ws, i, 5,  turno.fecha,                bg=bg)
        _data_cell(ws, i, 6,  turno.turno,                bg=bg)
        _data_cell(ws, i, 7,  turno.nombre_empleado,      bg=bg, align="left")
        _data_cell(ws, i, 8,  m["total_producido"],       bg=bg)
        _data_cell(ws, i, 9,  m["produccion_teorica"],    bg=bg)
        _data_cell(ws, i, 10, m["piezas_buenas"],         bg=bg)
        _data_cell(ws, i, 11, m["total_desperdicio"],     bg=_color_oee(100 - m["tasa_desperdicio"]))
        _data_cell(ws, i, 12, m["total_paradas_min"],     bg=bg)
        _data_cell(ws, i, 13, m["num_paradas_no_prog"],   bg=bg)
        _data_cell(ws, i, 14, m["disponibilidad"],        bg=_color_oee(m["disponibilidad"]), num_format='0.0"%"')
        _data_cell(ws, i, 15, m["rendimiento"],           bg=_color_oee(m["rendimiento"]),    num_format='0.0"%"')
        _data_cell(ws, i, 16, m["calidad"],               bg=_color_oee(m["calidad"]),        num_format='0.0"%"')
        _data_cell(ws, i, 17, m["oee"],                   bg=_color_oee(m["oee"]),            num_format='0.0"%"', bold=True)
        _data_cell(ws, i, 18, m["mttr"],                  bg=bg)

    _set_col_widths(ws, {
        "A":12,"B":22,"C":14,"D":7,"E":11,"F":7,"G":18,
        "H":10,"I":10,"J":10,"K":10,"L":10,"M":10,
        "N":10,"O":10,"P":10,"Q":10,"R":10
    })
    ws.freeze_panes = "A4"
    return ws


def _hoja_detalle_turno(wb, turno, orden, metricas):
    ws = wb.create_sheet(f"Turno {turno.id}")
    info = [
        ("Orden",        orden.numero_orden),
        ("Producto",     orden.descripcion_producto),
        ("Material",     orden.material),
        ("Tipo máquina", orden.tipo_maquina),
        ("Máquina N°",   orden.numero_maquina),
        ("Fecha",        turno.fecha),
        ("Turno",        turno.turno),
        ("Empleado",     turno.nombre_empleado),
        ("Cédula",       turno.cedula_empleado),
        ("Hora inicio",  turno.hora_inicio),
        ("Hora fin",     turno.hora_fin or "En curso"),
        ("Líder",        orden.nombre_lider),
    ]
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"DETALLE — {orden.numero_orden} / {turno.fecha} / Turno {turno.turno}"
    c.fill = _fill(COLOR_TITULO_SHEET)
    c.font = _font(bold=True, color="FFFFFF", size=12)
    c.alignment = _center()
    ws.row_dimensions[1].height = 25

    for i, (label, val) in enumerate(info, 2):
        ws.cell(row=i, column=1, value=label).font  = _font(bold=True, size=10)
        ws.cell(row=i, column=1).fill   = _fill(COLOR_HEADER_LIGHT)
        ws.cell(row=i, column=1).border = _border_thin()
        ws.cell(row=i, column=2, value=val).border = _border_thin()
        ws.cell(row=i, column=2).font   = _font(size=10)

    fila_kpi = len(info) + 3
    ws.merge_cells(f"A{fila_kpi}:F{fila_kpi}")
    c = ws.cell(row=fila_kpi, column=1, value="KPIs LEAN MANUFACTURING")
    c.fill = _fill(COLOR_LEAN_ACCENT)
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _center()

    kpis = [
        ("Producción real",     metricas["total_producido"],    "uds"),
        ("Producción teórica",  metricas["produccion_teorica"], "uds"),
        ("Piezas buenas",       metricas["piezas_buenas"],      "uds"),
        ("Total desperdicio",   metricas["total_desperdicio"],  "uds"),
        ("Tasa de desperdicio", metricas["tasa_desperdicio"],   "%"),
        ("Tiempo operativo",    metricas["tiempo_operativo"],   "min"),
        ("Tiempo paradas",      metricas["total_paradas_min"],  "min"),
        ("Paradas no prog.",    metricas["num_paradas_no_prog"],"eventos"),
        ("DISPONIBILIDAD",      metricas["disponibilidad"],     "%"),
        ("RENDIMIENTO",         metricas["rendimiento"],        "%"),
        ("CALIDAD",             metricas["calidad"],            "%"),
        ("OEE",                 metricas["oee"],                "%"),
        ("UPH real",            metricas["uph_real"],           "uds/h"),
        ("UPH teórica",         metricas["uph_teorica"],        "uds/h"),
        ("MTTR",                metricas["mttr"],               "min"),
        ("MTBF",                metricas["mtbf"],               "min"),
    ]
    for j, (label, val, unidad) in enumerate(kpis):
        r = fila_kpi + 1 + j
        ws.cell(row=r, column=1, value=label).font  = _font(bold=True, size=10)
        ws.cell(row=r, column=1).fill   = _fill(COLOR_HEADER_LIGHT)
        ws.cell(row=r, column=1).border = _border_thin()
        c_val = ws.cell(row=r, column=2, value=val)
        c_val.font   = _font(bold=(label in ["OEE","DISPONIBILIDAD","RENDIMIENTO","CALIDAD"]), size=10)
        c_val.border = _border_thin()
        if label in ["OEE","DISPONIBILIDAD","RENDIMIENTO","CALIDAD"]:
            c_val.fill = _fill(_color_oee(val))
        ws.cell(row=r, column=3, value=unidad).font   = _font(size=9, color="666666")
        ws.cell(row=r, column=3).border = _border_thin()

    _set_col_widths(ws, {"A":22, "B":14, "C":10, "D":14, "E":16, "F":14})
    ws.freeze_panes = "A2"


def _hoja_pareto_paradas(wb, turnos):
    ws = wb.create_sheet("Pareto Paradas")
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ANÁLISIS PARETO DE PARADAS — Lean Manufacturing"
    c.fill = _fill(COLOR_TITULO_SHEET)
    c.font = _font(bold=True, color="FFFFFF", size=12)
    c.alignment = _center()

    paradas_agrup: dict = {}
    for turno in turnos:
        for p in turno.paradas:
            key = f"{p.codigo} - {p.descripcion}"
            if key not in paradas_agrup:
                paradas_agrup[key] = {"minutos": 0, "ocurrencias": 0, "programada": p.programada}
            paradas_agrup[key]["minutos"]     += p.minutos
            paradas_agrup[key]["ocurrencias"] += 1

    sorted_paradas = sorted(paradas_agrup.items(), key=lambda x: x[1]["minutos"], reverse=True)
    total_min = sum(v["minutos"] for _, v in sorted_paradas)

    for col, h in enumerate(["Parada","Ocurrencias","Minutos totales","% del total","% Acumulado","Tipo"], 1):
        _header_cell(ws, 2, col, h, bg=COLOR_HEADER_DARK)

    acumulado = 0
    for i, (key, data) in enumerate(sorted_paradas, 3):
        pct      = round(data["minutos"] / total_min * 100, 1) if total_min > 0 else 0
        acumulado += pct
        tipo     = "Programada" if data["programada"] else "No programada"
        bg_acum  = COLOR_VERDE_GOOD if acumulado <= 80 else COLOR_AMARILLO_WARN
        _data_cell(ws, i, 1, key,                 align="left")
        _data_cell(ws, i, 2, data["ocurrencias"])
        _data_cell(ws, i, 3, data["minutos"])
        _data_cell(ws, i, 4, pct,                 num_format='0.0"%"')
        _data_cell(ws, i, 5, round(acumulado, 1), bg=bg_acum, num_format='0.0"%"')
        _data_cell(ws, i, 6, tipo,                bg=COLOR_ROJO_BAD if not data["programada"] else COLOR_VERDE_GOOD)

    _set_col_widths(ws, {"A":35,"B":13,"C":15,"D":12,"E":13,"F":14})


def _hoja_desperdicios(wb, turnos):
    ws = wb.create_sheet("Análisis Desperdicios")
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "ANÁLISIS DE DESPERDICIOS — Lean Manufacturing"
    c.fill = _fill(COLOR_TITULO_SHEET)
    c.font = _font(bold=True, color="FFFFFF", size=12)
    c.alignment = _center()

    des_agrup: dict = {}
    total_prod = 0
    for turno in turnos:
        total_prod += sum(r.cantidad for r in turno.registros_produccion)
        for d in turno.desperdicios:
            if d.defecto not in des_agrup:
                des_agrup[d.defecto] = {"cantidad": 0, "ocurrencias": 0}
            des_agrup[d.defecto]["cantidad"]    += d.cantidad
            des_agrup[d.defecto]["ocurrencias"] += 1

    sorted_des = sorted(des_agrup.items(), key=lambda x: x[1]["cantidad"], reverse=True)
    total_des  = sum(v["cantidad"] for _, v in sorted_des)

    for col, h in enumerate(["Defecto","Ocurrencias","Cantidad total","% del desperdicio","% de producción"], 1):
        _header_cell(ws, 2, col, h, bg=COLOR_HEADER_DARK)

    for i, (defecto, data) in enumerate(sorted_des, 3):
        pct_des  = round(data["cantidad"] / total_des  * 100, 1) if total_des  > 0 else 0
        pct_prod = round(data["cantidad"] / total_prod * 100, 2) if total_prod > 0 else 0
        _data_cell(ws, i, 1, defecto,  align="left")
        _data_cell(ws, i, 2, data["ocurrencias"])
        _data_cell(ws, i, 3, data["cantidad"])
        _data_cell(ws, i, 4, pct_des,  num_format='0.0"%"')
        _data_cell(ws, i, 5, pct_prod, bg=COLOR_ROJO_BAD if pct_prod > 2 else COLOR_VERDE_GOOD, num_format='0.00"%"')

    _set_col_widths(ws, {"A":30,"B":13,"C":15,"D":18,"E":18})


def _generar_libro(turnos, db, titulo) -> BytesIO:
    if not turnos:
        raise HTTPException(status_code=404, detail="No hay datos para exportar")

    orden_ids   = list({t.orden_id for t in turnos})
    ordenes     = db.query(Orden).filter(Orden.id.in_(orden_ids)).all()
    ordenes_map = {o.id: o for o in ordenes}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _hoja_resumen_lean(wb, turnos, ordenes_map, titulo)
    for turno in turnos:
        orden = ordenes_map.get(turno.orden_id)
        if orden:
            m = calcular_metricas_lean(turno, orden)
            _hoja_detalle_turno(wb, turno, orden, m)
    _hoja_pareto_paradas(wb, turnos)
    _hoja_desperdicios(wb, turnos)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _response_excel(output, filename):
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ─────────────────────────────────────────────
# HELPERS OEE PARA ENDPOINTS JSON
# ─────────────────────────────────────────────

def _calcular_oee_turno(turno: Turno, orden: Orden) -> dict:
    contador  = sum(r.cantidad for r in turno.registros_produccion)
    paradas   = turno.paradas
    min_np    = sum(p.minutos for p in paradas if not p.programada)

    try:
        def parse_hora(h: str) -> float:
            if not h:
                return 0.0
            clean = h.replace(".", "").replace("  ", " ").strip().lower()
            import re
            m = re.match(r"(\d{1,2}):(\d{2})(?:\s*(am|pm))?", clean)
            if not m:
                return 0.0
            hh, mm, ampm = int(m.group(1)), int(m.group(2)), m.group(3)
            if ampm == "pm" and hh != 12:
                hh += 12
            if ampm == "am" and hh == 12:
                hh = 0
            return hh + mm / 60

        inicio = parse_hora(turno.hora_inicio)
        fin    = parse_hora(turno.hora_fin) if turno.hora_fin else (
            datetime.now().hour + datetime.now().minute / 60
        )
        tiempo_real = fin - inicio
        if tiempo_real < 0:
            tiempo_real += 24
        tiempo_real = min(tiempo_real, 12)
    except Exception:
        tiempo_real = 0

    tiempo_programado = 12.0
    ciclo             = float(orden.ciclos or 0)
    cavidades         = int(orden.cavidades or 1)
    tiempo_disp       = max(tiempo_real - min_np / 60, 0)
    disponibilidad    = (tiempo_disp / tiempo_programado * 100) if tiempo_programado > 0 else 0
    prod_planeada     = (tiempo_programado * 3600 / ciclo * cavidades) if ciclo > 0 else 0
    rendimiento       = (contador / prod_planeada * 100) if prod_planeada > 0 else 0
    calidad           = 100.0
    oee               = (disponibilidad / 100) * (rendimiento / 100) * (calidad / 100) * 100

    return {
        "oee":            round(min(oee, 150), 1),
        "disponibilidad": round(min(disponibilidad, 150), 1),
        "rendimiento":    round(min(rendimiento, 150), 1),
        "calidad":        calidad,
    }


# ─────────────────────────────────────────────
# ENDPOINTS EXCEL
# ─────────────────────────────────────────────

@router.get("/reporte/turno/{turno_id}")
def reporte_por_turno(turno_id: int, db: Session = Depends(get_db)):
    turno = db.query(Turno).filter(Turno.id == turno_id).first()
    if not turno:
        raise HTTPException(status_code=404, detail="Turno no encontrado")
    titulo   = f"Reporte Turno {turno_id} — {turno.fecha}"
    output   = _generar_libro([turno], db, titulo)
    filename = f"reporte_turno_{turno_id}_{turno.fecha}.xlsx"
    return _response_excel(output, filename)


@router.get("/reporte/orden/{orden_id}")
def reporte_por_orden(orden_id: int, db: Session = Depends(get_db)):
    orden = db.query(Orden).filter(Orden.id == orden_id).first()
    if not orden:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    turnos   = db.query(Turno).filter(Turno.orden_id == orden_id).all()
    titulo   = f"Reporte Orden {orden.numero_orden} — {orden.descripcion_producto}"
    output   = _generar_libro(turnos, db, titulo)
    filename = f"reporte_orden_{orden.numero_orden}.xlsx"
    return _response_excel(output, filename)


@router.get("/reporte/fechas")
def reporte_por_fechas(
    fecha_inicio: str,
    fecha_fin:    str,
    db: Session = Depends(get_db)
):
    turnos = db.query(Turno).filter(
        Turno.fecha >= fecha_inicio,
        Turno.fecha <= fecha_fin
    ).all()
    if not turnos:
        raise HTTPException(status_code=404, detail="No hay turnos en ese rango de fechas")
    titulo   = f"Reporte Lean — {fecha_inicio} al {fecha_fin}"
    output   = _generar_libro(turnos, db, titulo)
    filename = f"reporte_{fecha_inicio}_{fecha_fin}.xlsx"
    return _response_excel(output, filename)


# ─────────────────────────────────────────────
# ENDPOINT JSON OEE PARA DASHBOARD
# ─────────────────────────────────────────────

@router.get("/oee/fechas")
def oee_por_fechas_json(
    fecha_inicio: str,
    fecha_fin:    str,
    db: Session = Depends(get_db)
):
    turnos = db.query(Turno).filter(
        Turno.fecha >= fecha_inicio,
        Turno.fecha <= fecha_fin
    ).all()

    if not turnos:
        return []

    orden_ids   = list({t.orden_id for t in turnos})
    ordenes     = db.query(Orden).filter(Orden.id.in_(orden_ids)).all()
    ordenes_map = {o.id: o for o in ordenes}

    resultado = []
    for turno in turnos:
        orden = ordenes_map.get(turno.orden_id)
        if not orden:
            continue
        kpis = _calcular_oee_turno(turno, orden)
        resultado.append({
            "turno_id":       turno.id,
            "orden_id":       turno.orden_id,
            "fecha":          turno.fecha,
            "turno":          turno.turno,
            "tipo_maquina":   orden.tipo_maquina,
            "numero_maquina": orden.numero_maquina,
            "oee":            kpis["oee"],
            "disponibilidad": kpis["disponibilidad"],
            "rendimiento":    kpis["rendimiento"],
            "calidad":        kpis["calidad"],
        })

    return resultado


# ─────────────────────────────────────────────
# ENDPOINT EMPLEADOS MES
# ─────────────────────────────────────────────

@router.get("/empleados-mes")
def reporte_empleados_mes(mes: int, anio: int, db: Session = Depends(get_db)):
    primer_dia = date(anio, mes, 1)
    ultimo_dia = date(anio, mes, calendar.monthrange(anio, mes)[1])

    turnos = db.query(Turno).filter(
        Turno.fecha >= str(primer_dia),
        Turno.fecha <= str(ultimo_dia),
        Turno.hora_fin != None
    ).all()

    empleados = {}
    for t in turnos:
        ced   = t.cedula_empleado
        orden = db.query(Orden).filter(Orden.id == t.orden_id).first()
        if not orden:
            continue
        kpis     = _calcular_oee_turno(t, orden)
        contador = sum(r.cantidad for r in t.registros_produccion)

        if ced not in empleados:
            empleados[ced] = {
                "cedula": ced, "nombre": t.nombre_empleado,
                "turnos": 0, "oee_sum": 0.0, "disp_sum": 0.0,
                "rend_sum": 0.0, "total_produccion": 0,
            }
        empleados[ced]["turnos"]           += 1
        empleados[ced]["oee_sum"]          += kpis["oee"]
        empleados[ced]["disp_sum"]         += kpis["disponibilidad"]
        empleados[ced]["rend_sum"]         += kpis["rendimiento"]
        empleados[ced]["total_produccion"] += contador

    resultado = []
    for emp in empleados.values():
        n        = emp["turnos"]
        oee_prom = emp["oee_sum"] / n if n > 0 else 0
        resultado.append({
            "cedula":                  emp["cedula"],
            "nombre":                  emp["nombre"],
            "turnos":                  n,
            "oee_promedio":            round(oee_prom, 2),
            "disponibilidad_promedio": round(emp["disp_sum"] / n, 2) if n > 0 else 0,
            "rendimiento_promedio":    round(emp["rend_sum"] / n, 2) if n > 0 else 0,
            "total_produccion":        emp["total_produccion"],
            "bono":                    oee_prom >= 94.0,
        })

    return sorted(resultado, key=lambda x: x["oee_promedio"], reverse=True)


# ─────────────────────────────────────────────
# ENDPOINT CONSUMO
# ─────────────────────────────────────────────

@router.get("/consumo/datos")
def get_datos_consumo(
    fecha_inicio: str,
    fecha_fin: str,
    db: Session = Depends(get_db)
):
    """Retorna datos de consumo por máquina en un rango de fechas."""
    from datetime import datetime as dt

    try:
        fi = dt.strptime(fecha_inicio, "%Y-%m-%d")
        ff = dt.strptime(fecha_fin, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato inválido. Usar YYYY-MM-DD")

    turnos = (
        db.query(Turno)
        .join(Orden, Turno.orden_id == Orden.id)
        .filter(Turno.fecha >= fi, Turno.fecha <= ff)
        .all()
    )

    resultado = {}

    for turno in turnos:
        orden = db.query(Orden).filter(Orden.id == turno.orden_id).first()
        if not orden:
            continue

        tipo = orden.tipo_maquina
        num  = str(orden.numero_maquina)
        key  = f"{tipo}_{num}"

        if key not in resultado:
            resultado[key] = {
                "tipo_maquina":       tipo,
                "numero_maquina":     num,
                "horas_trabajadas":   0.0,
                "kwh_consumidos":     0.0,
                "kg_procesados":      0.0,
                "unidades_producidas": 0,
            }

        # Horas trabajadas
        horas = 0.0
        if turno.hora_inicio and turno.hora_fin:
            try:
                hi   = dt.strptime(turno.hora_inicio[:5], "%H:%M")
                hf   = dt.strptime(turno.hora_fin[:5],    "%H:%M")
                diff = (hf - hi).seconds / 3600
                horas = diff if diff > 0 else 0.0
            except Exception:
                horas = 0.0

        resultado[key]["horas_trabajadas"] = round(
            resultado[key]["horas_trabajadas"] + horas, 2
        )

        # kWh
        kw = KW_MAQUINA.get((tipo, num), 0.0)
        resultado[key]["kwh_consumidos"] = round(
            resultado[key]["kwh_consumidos"] + kw * horas, 2
        )

        # Unidades producidas
        registros      = db.query(RegistroProduccion).filter(
            RegistroProduccion.turno_id == turno.id
        ).all()
        total_unidades = sum(r.cantidad for r in registros)
        resultado[key]["unidades_producidas"] += total_unidades

        # kg procesados (solo si el producto tiene peso_pieza)
        producto = db.query(Producto).filter(
            Producto.codigo == orden.codigo_producto
        ).first()
        if producto and producto.peso_pieza:
            cavidades = orden.cavidades or 1
            kg = (total_unidades * cavidades * producto.peso_pieza) / 1000
            resultado[key]["kg_procesados"] = round(
                resultado[key]["kg_procesados"] + kg, 2
            )

    return list(resultado.values())

# ─────────────────────────────────────────────
# ENDPOINT MANTENIMIENTO
# ─────────────────────────────────────────────

@router.get("/mantenimiento/datos")
def get_datos_mantenimiento(
    fecha_inicio: str,
    fecha_fin:    str,
    db: Session = Depends(get_db)
):
    """
    Retorna MTTR, MTBF, disponibilidad y top causas por máquina
    en un rango de fechas. También incluye tendencia mensual.
    """
    from datetime import datetime as dt

    try:
        fi = dt.strptime(fecha_inicio, "%Y-%m-%d")
        ff = dt.strptime(fecha_fin,    "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato inválido. Usar YYYY-MM-DD")

    turnos = (
        db.query(Turno)
        .join(Orden, Turno.orden_id == Orden.id)
        .filter(Turno.fecha >= fecha_inicio, Turno.fecha <= fecha_fin)
        .all()
    )

    # ── Acumular datos por máquina ────────────────────────────────────────────
    maquinas: dict = {}

    for turno in turnos:
        orden = db.query(Orden).filter(Orden.id == turno.orden_id).first()
        if not orden:
            continue

        tipo = orden.tipo_maquina
        num  = str(orden.numero_maquina)

        # Determinar tipo_key y label
        if tipo == "linea" and num == "1":
            tipo_key = "linea_copro"
            label    = "L. Copro"
        elif tipo == "linea" and num == "2":
            tipo_key = "linea_orina"
            label    = "L. Orina"
        elif tipo == "inyeccion":
            tipo_key = f"inyeccion_{num}"
            label    = f"Iny. {num}"
        elif tipo == "acondicionamiento":
            tipo_key = f"acondicionamiento_{num}"
            label    = f"Acond. {num}"
        else:
            tipo_key = f"{tipo}_{num}"
            label    = f"{tipo.capitalize()} {num}"

        if tipo_key not in maquinas:
            maquinas[tipo_key] = {
                "tipo_key":       tipo_key,
                "tipo_maquina":   tipo,
                "numero_maquina": num,
                "label":          label,
                "color":          _color_tipo(tipo),
                "min_np_total":   0,
                "n_paradas":      0,
                "tiempo_h_total": 0.0,
                "causas":         {},
                "por_mes":        {},  # mes -> {min_np, n_paradas, tiempo_h}
            }

        maq = maquinas[tipo_key]

        # Calcular horas del turno
        horas = _horas_turno(turno)
        maq["tiempo_h_total"] += horas

        # Paradas no programadas
        paradas_np = [p for p in turno.paradas if not p.programada]
        min_np     = sum(p.minutos for p in paradas_np)
        maq["min_np_total"] += min_np
        maq["n_paradas"]    += len(paradas_np)

        # Causas acumuladas
        for p in paradas_np:
            desc = p.descripcion
            if desc not in maq["causas"]:
                maq["causas"][desc] = {"minutos": 0, "ocurrencias": 0}
            maq["causas"][desc]["minutos"]     += p.minutos
            maq["causas"][desc]["ocurrencias"] += 1

        # Tendencia mensual
        try:
            mes_num = int(turno.fecha[5:7])
        except:
            mes_num = fi.month

        if mes_num not in maq["por_mes"]:
            maq["por_mes"][mes_num] = {"min_np": 0, "n": 0, "h": 0.0}
        maq["por_mes"][mes_num]["min_np"] += min_np
        maq["por_mes"][mes_num]["n"]      += len(paradas_np)
        maq["por_mes"][mes_num]["h"]      += horas

    # ── Calcular métricas finales ─────────────────────────────────────────────
    resultado = []
    for maq in maquinas.values():
        n      = maq["n_paradas"]
        h      = maq["tiempo_h_total"]
        min_np = maq["min_np_total"]
        tiempo_op_min = h * 60 - min_np

        mttr = round(min_np / n, 1)             if n > 0 else 0
        mtbf = round(tiempo_op_min / n, 1)      if n > 0 else round(h * 60, 1)
        disp = round(tiempo_op_min / (h*60)*100, 1) if h > 0 else 100.0

        # Top 5 causas
        top_causas = sorted(
            [{"descripcion": k, **v} for k, v in maq["causas"].items()],
            key=lambda x: x["minutos"], reverse=True
        )[:5]

        # Tendencia mensual
        tendencia = []
        for mes_num, d in sorted(maq["por_mes"].items()):
            n_m   = d["n"]
            h_m   = d["h"]
            mn_m  = d["min_np"]
            top_m = h_m * 60 - mn_m
            tendencia.append({
                "mes":  mes_num,
                "mttr": round(mn_m / n_m, 1)         if n_m > 0 else 0,
                "mtbf": round(top_m / n_m, 1)         if n_m > 0 else round(h_m*60, 1),
                "disp": round(top_m / (h_m*60)*100, 1) if h_m > 0 else 100.0,
            })

        resultado.append({
            "tipo_key":       maq["tipo_key"],
            "tipo_maquina":   maq["tipo_maquina"],
            "numero_maquina": maq["numero_maquina"],
            "label":          maq["label"],
            "color":          maq["color"],
            "mttr":           mttr,
            "mtbf":           mtbf,
            "disponibilidad": disp,
            "n_paradas":      n,
            "min_paradas":    min_np,
            "top_causas":     top_causas,
            "tendencia":      tendencia,
        })

    # Ordenar por disponibilidad ascendente (peores primero)
    resultado.sort(key=lambda x: x["disponibilidad"])
    return resultado


def _color_tipo(tipo: str) -> str:
    return {
        "inyeccion":         "#378add",
        "soplado":           "#639922",
        "linea":             "#ef9f27",
        "acondicionamiento": "#1D9E75",
    }.get(tipo, "#888888")


def _horas_turno(turno) -> float:
    from datetime import datetime as dt
    try:
        def ph(h):
            if not h: return 0.0
            import re
            clean = h.replace(".", "").replace("  ", " ").strip().lower()
            m = re.match(r"(\d{1,2}):(\d{2})(?:\s*(am|pm))?", clean)
            if not m: return 0.0
            hh, mm, ampm = int(m.group(1)), int(m.group(2)), m.group(3)
            if ampm == "pm" and hh != 12: hh += 12
            if ampm == "am" and hh == 12: hh = 0
            return hh + mm / 60
        inicio = ph(turno.hora_inicio)
        fin    = ph(turno.hora_fin) if turno.hora_fin else (
            dt.now().hour + dt.now().minute / 60
        )
        diff = fin - inicio
        if diff < 0: diff += 24
        return min(diff, 12)
    except:
        return 0.0    